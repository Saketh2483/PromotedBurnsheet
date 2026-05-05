"""agent_tools.py - Strands Agent tools for timesheet reconciliation"""
import json
from pathlib import Path
import openpyxl
from sow_config import determine_sow_stream, get_daily_hours

try:
    from strands import Agent, tool
    from strands.models.bedrock import BedrockModel
except Exception:
    Agent = None
    def tool(fn=None, **kw):
        if fn is not None:
            return fn
        return lambda f: f
    class BedrockModel:
        def __init__(self, *a, **kw): raise RuntimeError("strands-agents not installed")

EXCEL_PATH = Path(__file__).resolve().parent.parent / "public" / "Combined-Input.xlsx"


@tool
def reconcile_timesheet(identifier: str, new_timesheet_hrs: float, update_excel: bool = True) -> str:
    """Update timesheet hours for an employee. Recalculates Actual Rate & Variance."""
    if not EXCEL_PATH.exists():
        return json.dumps({"error": f"File not found - {EXCEL_PATH}"})
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}
    for r in ["EMPId", "Name", "Timesheet", "Hourly Rate($)", "Projected Rate($)", "Actual Rate", "Variance"]:
        if r not in col:
            wb.close()
            return json.dumps({"error": f"Missing column: {r}"})

    search = identifier.strip().lower()
    updates = []
    for row in range(2, ws.max_row + 1):
        raw_id = str(ws.cell(row=row, column=col["EMPId"]).value or "").strip()
        emp_id = raw_id[:-2] if raw_id.endswith(".0") else raw_id
        name = str(ws.cell(row=row, column=col["Name"]).value or "").strip()
        if search != emp_id.lower() and search not in name.lower():
            continue
        old = float(ws.cell(row=row, column=col["Timesheet"]).value or 0)
        rate = float(ws.cell(row=row, column=col["Hourly Rate($)"]).value or 0)
        proj = float(ws.cell(row=row, column=col["Projected Rate($)"]).value or 0)
        loc = str(ws.cell(row=row, column=col.get("Location", 0)).value or "") if "Location" in col else ""
        ctry = str(ws.cell(row=row, column=col.get("Country", 0)).value or "") if "Country" in col else ""
        ws.cell(row=row, column=col["Timesheet"]).value = new_timesheet_hrs
        actual = round(rate * new_timesheet_hrs, 2)
        variance = round(actual - proj, 2)
        ws.cell(row=row, column=col["Actual Rate"]).value = actual
        ws.cell(row=row, column=col["Variance"]).value = variance
        sow = determine_sow_stream(loc, ctry)
        updates.append({"row": row, "empId": emp_id, "name": name, "oldTimesheetHrs": old,
                        "newTimesheetHrs": new_timesheet_hrs, "rateUsd": rate, "projectedRate": proj,
                        "newActualRate": actual, "newVariance": variance, "sowStream": sow,
                        "dailyHours": get_daily_hours(sow), "location": loc, "country": ctry})
    if not updates:
        wb.close()
        return json.dumps({"error": f"No employee found matching '{identifier}'."})
    if update_excel:
        wb.save(EXCEL_PATH)
    wb.close()
    return json.dumps({"status": "success", "excelUpdated": update_excel, "updates": updates})


@tool
def read_employee_data(identifier: str = "") -> str:
    """Read employee data from Excel."""
    if not EXCEL_PATH.exists():
        return json.dumps({"error": f"File not found - {EXCEL_PATH}"})
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
    if not identifier.strip():
        wb.close()
        return json.dumps({"total_rows": ws.max_row - 1, "columns": list(col_map.keys())})
    search = identifier.strip().lower()
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        eid = str(row[col_map.get("EMPId", 0)] or "").strip().lower()
        ename = str(row[col_map.get("Name", 0)] or "").strip().lower()
        if search == eid or search in ename:
            results.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
    wb.close()
    return json.dumps({"matches": len(results), "data": results[:10]}, default=str)


def get_employee_sow_info(identifier):
    """Get employee SOW stream and current timesheet for leave calculation."""
    if not EXCEL_PATH.exists():
        return None
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {str(h).strip(): i for i, h in enumerate(headers) if h}
    search = identifier.strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        eid = str(row[col.get("EMPId", 0)] or "").strip()
        if eid.endswith(".0"): eid = eid[:-2]
        ename = str(row[col.get("Name", 0)] or "").strip()
        if search == eid.lower() or search in ename.lower():
            loc = str(row[col.get("Location", 0)] or "") if "Location" in col else ""
            ctry = str(row[col.get("Country", 0)] or "") if "Country" in col else ""
            ts = float(row[col.get("Timesheet", 0)] or 0) if "Timesheet" in col else 0
            rate = float(row[col.get("Hourly Rate($)", 0)] or 0) if "Hourly Rate($)" in col else 0
            proj = float(row[col.get("Projected Rate($)", 0)] or 0) if "Projected Rate($)" in col else 0
            sow = determine_sow_stream(loc, ctry)
            wb.close()
            return {"empId": eid, "name": ename, "sowStream": sow, "dailyHours": get_daily_hours(sow),
                    "currentTimesheet": ts, "rateUsd": rate, "projectedRate": proj}
    wb.close()
    return None
