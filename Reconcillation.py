"""
Reconciliation Agent - AWS Strands + Bedrock + FastAPI
======================================================
Exposes a FastAPI server that lets calling clients trigger a
reconciliation agent powered by AWS Bedrock (Claude) via the
Strands Agents SDK.

Environment variables required (loaded from .env):
  - AWS_BEARER_TOKEN_BEDROCK : pre-signed bearer token for Bedrock
  - AWS_REGION               : AWS region (default: us-west-1)

Start:
    python Reconcillation.py
"""

import os, sys, json, base64, traceback, re
import importlib
from typing import Optional, List, Dict
from urllib.parse import parse_qs, urlparse, unquote
from datetime import datetime, timedelta, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# 1. Environment & credential bootstrap
# ---------------------------------------------------------------------------
try:
    _dotenv = importlib.import_module("dotenv")
    load_dotenv = getattr(_dotenv, "load_dotenv", lambda: None)
    load_dotenv()
except Exception:
    def load_dotenv():
        return None

BEARER_TOKEN = os.getenv("AWS_BEARER_TOKEN_BEDROCK", "")
REGION = os.getenv("AWS_REGION", "us-west-1")

EXCEL_PATH = Path(__file__).resolve().parent / "public" / "Combined-Input.xlsx"

REQUIRED_ENV_VARS = ["AWS_BEARER_TOKEN_BEDROCK", "AWS_REGION"]

# ---------------------------------------------------------------------------
# Pending state: stores updates not yet reconciled to Excel
# Key: employee identifier (lowercased), Value: dict with update details
# ---------------------------------------------------------------------------
_pending_updates: Dict[str, dict] = {}


def _extract_credentials_from_bearer_token():
    """Decode the pre-signed-URL bearer token and set temporary AWS creds."""
    encoded = BEARER_TOKEN
    if encoded.startswith("bedrock-api-key-"):
        encoded = encoded[len("bedrock-api-key-"):]
    padding = 4 - len(encoded) % 4
    if padding != 4:
        encoded += "=" * padding
    decoded_url = base64.b64decode(encoded).decode()
    if not decoded_url.startswith("http"):
        decoded_url = "https://" + decoded_url
    parsed = urlparse(decoded_url)
    params = parse_qs(parsed.query)
    credential = params.get("X-Amz-Credential", [""])[0]
    cred_parts = credential.split("/")
    access_key = cred_parts[0]
    cred_region = cred_parts[2]
    security_token = unquote(params.get("X-Amz-Security-Token", [""])[0])
    amz_date = params.get("X-Amz-Date", [""])[0]
    expires = int(params.get("X-Amz-Expires", ["0"])[0])

    token_time = datetime.strptime(amz_date, "%Y%m%dT%H%M%SZ").replace(tzinfo=timezone.utc)
    expiry_time = token_time + timedelta(seconds=expires)
    now = datetime.now(timezone.utc)
    if now > expiry_time:
        raise RuntimeError(f"Bearer token expired at {expiry_time} (now={now})")

    os.environ["AWS_ACCESS_KEY_ID"] = access_key
    os.environ["AWS_SECRET_ACCESS_KEY"] = access_key
    os.environ["AWS_SESSION_TOKEN"] = security_token
    os.environ["AWS_DEFAULT_REGION"] = cred_region
    return {
        "access_key": access_key,
        "region": cred_region,
        "expires": str(expiry_time),
        "remaining": str(expiry_time - now),
    }


# ---------------------------------------------------------------------------
# 2. Strands imports
# ---------------------------------------------------------------------------
import openpyxl

try:
    from strands import Agent, tool
    from strands.models.bedrock import BedrockModel
    _STRANDS_AVAILABLE = True
except Exception:
    _STRANDS_AVAILABLE = False
    Agent = None

    def tool(fn=None, **kwargs):
        return fn

    class BedrockModel:
        def __init__(self, *args, **kwargs):
            raise RuntimeError("strands-agents not installed")


# ---------------------------------------------------------------------------
# 3. Tools
# ---------------------------------------------------------------------------

@tool
def reconcile_timesheet(
    identifier: str,
    new_timesheet_hrs: float,
    update_excel: bool = True,
) -> str:
    """Update Timesheet Hrs for a specific employee.

    If update_excel is True, writes changes to the Excel file.
    Otherwise returns computed values only (UI-only preview).

    Args:
        identifier: Employee ID or name to search for.
        new_timesheet_hrs: The new timesheet hours value to set.
        update_excel: Whether to persist changes to the Excel file.
    """
    fp = EXCEL_PATH
    if not fp.exists():
        return json.dumps({"error": f"File not found - {fp}"})

    wb = openpyxl.load_workbook(fp)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_map = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}

    required = ["EMPId", "Name", "Timesheet", "Hourly Rate($)", "Projected Rate($)",
                "Actual Rate", "Variance"]
    missing_cols = [c for c in required if c not in col_map]
    if missing_cols:
        wb.close()
        return json.dumps({"error": f"Missing columns: {missing_cols}"})

    id_col = col_map["EMPId"]
    name_col = col_map["Name"]
    ts_col = col_map["Timesheet"]
    rate_usd_col = col_map["Hourly Rate($)"]
    proj_col = col_map["Projected Rate($)"]
    act_col = col_map["Actual Rate"]
    var_col = col_map["Variance"]

    search = identifier.strip().lower()
    matched_rows = []
    for row_num in range(2, ws.max_row + 1):
        raw_id = str(ws.cell(row=row_num, column=id_col).value or "").strip()
        emp_id = raw_id[:-2] if raw_id.endswith(".0") else raw_id
        emp_name = str(ws.cell(row=row_num, column=name_col).value or "").strip()
        if search == emp_id.lower() or search in emp_name.lower():
            matched_rows.append(row_num)

    if not matched_rows:
        wb.close()
        return json.dumps({"error": f"No employee found matching '{identifier}'."})

    updates = []
    for row_num in matched_rows:
        old_hrs = ws.cell(row=row_num, column=ts_col).value or 0
        rate_usd = float(ws.cell(row=row_num, column=rate_usd_col).value or 0)
        projected = float(ws.cell(row=row_num, column=proj_col).value or 0)

        ws.cell(row=row_num, column=ts_col).value = new_timesheet_hrs
        new_actual = round(rate_usd * new_timesheet_hrs, 2)
        new_variance = round(new_actual - projected, 2)
        ws.cell(row=row_num, column=act_col).value = new_actual
        ws.cell(row=row_num, column=var_col).value = new_variance

        emp_id_raw = str(ws.cell(row=row_num, column=id_col).value or "").strip()
        if emp_id_raw.endswith(".0"):
            emp_id_raw = emp_id_raw[:-2]

        updates.append({
            "row": row_num,
            "empId": emp_id_raw,
            "name": str(ws.cell(row=row_num, column=name_col).value or "").strip(),
            "oldTimesheetHrs": float(old_hrs),
            "newTimesheetHrs": new_timesheet_hrs,
            "rateUsd": rate_usd,
            "projectedRate": projected,
            "newActualRate": new_actual,
            "newVariance": new_variance,
        })

    if update_excel:
        wb.save(fp)
    wb.close()

    return json.dumps({
        "status": "success",
        "excelUpdated": update_excel,
        "message": f"{'Reconciled' if update_excel else 'Computed'} {len(updates)} row(s).",
        "updates": updates,
    })


@tool
def read_employee_data(identifier: str = "") -> str:
    """Read employee data from the Excel burnsheet.

    Args:
        identifier: Optional employee ID or name to search for.
    """
    fp = EXCEL_PATH
    if not fp.exists():
        return json.dumps({"error": f"File not found - {fp}"})

    wb = openpyxl.load_workbook(fp, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}

    if not identifier.strip():
        wb.close()
        return json.dumps({"total_rows": ws.max_row - 1, "columns": list(col_map.keys())})

    id_idx = col_map.get("EMPId")
    name_idx = col_map.get("Name")
    if id_idx is None or name_idx is None:
        wb.close()
        return json.dumps({"error": "Missing EMPId or Name column."})

    search = identifier.strip().lower()
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        emp_id = str(row[id_idx] or "").strip().lower()
        emp_name = str(row[name_idx] or "").strip().lower()
        if search == emp_id or search in emp_name:
            results.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})

    wb.close()
    return json.dumps({"matches": len(results), "data": results[:10]}, default=str)


# ---------------------------------------------------------------------------
# 4. Strands Agent (lazy init)
# ---------------------------------------------------------------------------
_agent = None


def _get_agent():
    global _agent
    if _agent is None:
        creds = _extract_credentials_from_bearer_token()
        model = BedrockModel(
            model_id="us.anthropic.claude-sonnet-4-20250514-v1:0",
            region_name=creds["region"],
        )
        _agent = Agent(
            model=model,
            tools=[reconcile_timesheet, read_employee_data],
            system_prompt=(
                "You are a Reconciliation Agent for a timesheet burnsheet system.\n"
                "The Excel file has columns: ESA ID, ESA Description, Verizon TQ ID, "
                "Verizon TQ Description, POC, EMPId, Name, Location, Country, ACT/PCT, "
                "Skill Set, Verizon Level Mapping, Classification, Key, Cognizant Designation, "
                "Service Line, Timesheet, Hourly Rate(Rs), Hourly Rate($), Projected Rate($), "
                "Actual Rate, Variance.\n\n"
                "Rules:\n"
                "1. Any timesheet update MUST have employee identifier AND hours value.\n"
                "2. Obey [SYSTEM OVERRIDE] instructions about update_excel flag.\n"
                "3. Always report old and new values clearly.\n"
            ),
        )
    return _agent


# ---------------------------------------------------------------------------
# 5. Helper functions for prompt parsing
# ---------------------------------------------------------------------------

def _extract_hours_from_prompt(text):
    """Extract numeric hours value from the prompt."""
    patterns = [
        r'\bto\s+(\d+(?:\.\d+)?)\b',
        r'(\d+(?:\.\d+)?)\s*(?:hours|hrs)\b',
        r'timesheet\s*(?:value|hours|hrs)?\s*(?:to|=|:)\s*(\d+(?:\.\d+)?)',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return float(m.group(1))
    return None


def _extract_identifier(text):
    """Extract employee identifier (empId or name) from the prompt."""
    # Numeric employee ID (4+ digits after keyword)
    id_match = re.search(r'(?:emp(?:loyee)?(?:\s*id)?[:\s]+|for\s+|of\s+)(\d{4,})', text, re.IGNORECASE)
    if id_match:
        return id_match.group(1)
    # Standalone large number (6+ digits, not hours)
    standalone = re.search(r'\b(\d{6,})\b', text)
    if standalone:
        return standalone.group(1)
    # Name pattern
    name_match = re.search(
        r'(?:for|of|employee)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)',
        text
    )
    if name_match:
        return name_match.group(1).strip()
    return None


# ---------------------------------------------------------------------------
# 6. FastAPI application
# ---------------------------------------------------------------------------
try:
    _fastapi = importlib.import_module("fastapi")
    FastAPI = getattr(_fastapi, "FastAPI")
    HTTPException = getattr(_fastapi, "HTTPException")
    _fastapi_cors = importlib.import_module("fastapi.middleware.cors")
    CORSMiddleware = getattr(_fastapi_cors, "CORSMiddleware")
    _pydantic = importlib.import_module("pydantic")
    BaseModel = getattr(_pydantic, "BaseModel")
    _FASTAPI_AVAILABLE = True
except Exception:
    _FASTAPI_AVAILABLE = False

    class HTTPException(Exception):
        def __init__(self, status_code=500, detail=""):
            super().__init__(detail)
            self.status_code = status_code
            self.detail = detail

    class BaseModel:
        def __init__(self, **kwargs):
            for k, v in kwargs.items():
                setattr(self, k, v)

    class FastAPI:
        def __init__(self, *a, **kw):
            pass

        def add_middleware(self, *a, **kw):
            pass

    class CORSMiddleware:
        pass

app = FastAPI(title="Reconciliation Agent API", version="1.0.0")
try:
    app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
except Exception:
    pass


class AgentRequest(BaseModel):
    prompt: str = ""


class AgentResponse(BaseModel):
    status: str
    response: str
    updates: Optional[List[dict]] = None
    excelUpdated: bool = False


class DirectReconcileRequest(BaseModel):
    identifier: str
    newTimesheetHrs: float


# --- Health ---
@app.get("/health")
def health_check():
    """Test connection to AWS Bedrock. Reports checked env vars on failure."""
    checked_vars = ["AWS_BEARER_TOKEN_BEDROCK", "AWS_REGION", "AWS_ACCESS_KEY_ID",
                    "AWS_SECRET_ACCESS_KEY", "AWS_DEFAULT_REGION", "AWS_SESSION_TOKEN"]
    missing = [v for v in REQUIRED_ENV_VARS if not os.getenv(v)]
    if missing:
        return {
            "status": "error",
            "message": f"Missing environment variables: {missing}",
            "checked_variables": checked_vars,
        }
    try:
        creds = _extract_credentials_from_bearer_token()
        return {"status": "ok", "credentials": creds, "checked_variables": checked_vars}
    except Exception as e:
        return {
            "status": "error",
            "message": str(e),
            "checked_variables": checked_vars,
        }


# --- Main reconcile endpoint ---
@app.post("/reconcile", response_model=AgentResponse)
def run_reconciliation(req: AgentRequest):
    """Handle reconciliation requests with Case 1/2/3 logic.

    Case 1: No 'reconcile' word -> UI only, store pending
    Case 2: 'reconcile' + employee + hours -> write Excel
    Case 3: Exactly 'reconcile' -> apply all pending updates to Excel
    """
    prompt = req.prompt.strip()
    prompt_lower = prompt.lower()

    # Test AWS connection
    checked_vars = ["AWS_BEARER_TOKEN_BEDROCK", "AWS_REGION", "AWS_ACCESS_KEY_ID",
                    "AWS_SECRET_ACCESS_KEY", "AWS_DEFAULT_REGION", "AWS_SESSION_TOKEN"]
    missing = [v for v in REQUIRED_ENV_VARS if not os.getenv(v)]
    if missing:
        raise HTTPException(
            status_code=503,
            detail=f"AWS connection failed. Missing env vars: {missing}. Checked: {checked_vars}"
        )
    try:
        _extract_credentials_from_bearer_token()
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"AWS connection failed: {str(e)}. Checked variables: {checked_vars}"
        )

    if not prompt:
        return AgentResponse(
            status="error",
            response="Please provide a message.",
            excelUpdated=False,
        )

    # ═══ CASE 3: Exactly "reconcile" -> apply only pending updates ═══
    if prompt_lower == "reconcile":
        try:
            if not _pending_updates:
                return AgentResponse(
                    status="no_pending",
                    response="No pending updates to reconcile. All data is already up to date.",
                    updates=None,
                    excelUpdated=False,
                )

            all_updates = []
            for key, pending in list(_pending_updates.items()):
                res_json = reconcile_timesheet(
                    identifier=pending["identifier"],
                    new_timesheet_hrs=pending["newTimesheetHrs"],
                    update_excel=True,
                )
                res = json.loads(res_json)
                if res.get("updates"):
                    all_updates.extend(res["updates"])

            _pending_updates.clear()

            if all_updates:
                return AgentResponse(
                    status="success",
                    response=f"Reconciliation complete. {len(all_updates)} row(s) updated in Excel.",
                    updates=all_updates,
                    excelUpdated=True,
                )
            else:
                return AgentResponse(
                    status="error",
                    response="Reconciliation failed. No rows were updated.",
                    excelUpdated=False,
                )
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=str(e))

    # Parse prompt for identifier and hours
    contains_reconcile = "reconcile" in prompt_lower
    extracted_hours = _extract_hours_from_prompt(prompt)
    identifier = _extract_identifier(prompt)

    # ═══ CASE 2: Contains 'reconcile' + identifier + hours -> write Excel ═══
    if contains_reconcile and identifier and extracted_hours is not None:
        try:
            result_json = reconcile_timesheet(
                identifier=identifier,
                new_timesheet_hrs=extracted_hours,
                update_excel=True,
            )
            result = json.loads(result_json)
            updates = result.get("updates")

            # Clear from pending if it was there
            if updates:
                for u in updates:
                    key = u.get("empId", "").lower()
                    _pending_updates.pop(key, None)

            return AgentResponse(
                status=result.get("status", "success"),
                response=f"Reconciled {len(updates)} row(s). Excel has been updated.",
                updates=updates,
                excelUpdated=True,
            )
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=str(e))

    # ═══ CASE 1 / CASE 2 with missing info -> UI only, store pending ═══
    if identifier and extracted_hours is not None:
        try:
            result_json = reconcile_timesheet(
                identifier=identifier,
                new_timesheet_hrs=extracted_hours,
                update_excel=False,
            )
            result = json.loads(result_json)
            updates = result.get("updates")

            if updates:
                for u in updates:
                    key = u.get("empId", identifier).lower()
                    _pending_updates[key] = {
                        "identifier": identifier,
                        "newTimesheetHrs": extracted_hours,
                        "empId": u.get("empId"),
                        "name": u.get("name"),
                        "oldTimesheetHrs": u.get("oldTimesheetHrs"),
                        "newActualRate": u.get("newActualRate"),
                        "newVariance": u.get("newVariance"),
                    }

                response_text = (
                    f"Timesheet update saved for review. Updated UI for {len(updates)} row(s). "
                    "Excel was NOT modified.\n\n"
                    "If you want to reconcile, then type reconcile."
                )
            else:
                response_text = result.get("error", "No matching employee found.")

            return AgentResponse(
                status=result.get("status", "error"),
                response=response_text,
                updates=updates,
                excelUpdated=False,
            )
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=str(e))

    # Missing identifier or hours — cannot process
    if contains_reconcile:
        msg = "Timesheet update saved for review. If you want to reconcile, then type reconcile."
    else:
        msg = ("I need both an employee identifier (empId or name) and a timesheet hours value "
               "to process an update. Please provide both.")

    return AgentResponse(
        status="error",
        response=msg,
        excelUpdated=False,
    )


# --- Direct reconcile endpoint (Case 2 shortcut) - always writes Excel ---
@app.post("/reconcile-direct")
def run_reconciliation_direct(req: DirectReconcileRequest):
    """Directly reconcile: validates and writes to Excel."""
    try:
        result_json = reconcile_timesheet(
            identifier=req.identifier,
            new_timesheet_hrs=req.newTimesheetHrs,
            update_excel=True,
        )
        result = json.loads(result_json)
        # Clear from pending
        if result.get("updates"):
            for u in result["updates"]:
                _pending_updates.pop(u.get("empId", "").lower(), None)
        return result
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# --- UI-only update (Case 1) - no Excel write, stores pending ---
@app.post("/update-ui")
def update_ui_only(req: DirectReconcileRequest):
    """Update timesheet for UI display only. Stores in pending state."""
    try:
        result_json = reconcile_timesheet(
            identifier=req.identifier,
            new_timesheet_hrs=req.newTimesheetHrs,
            update_excel=False,
        )
        result = json.loads(result_json)
        if result.get("status") == "success":
            updates = result.get("updates", [])
            for u in updates:
                key = u.get("empId", req.identifier).lower()
                _pending_updates[key] = {
                    "identifier": req.identifier,
                    "newTimesheetHrs": req.newTimesheetHrs,
                    "empId": u.get("empId"),
                    "name": u.get("name"),
                    "oldTimesheetHrs": u.get("oldTimesheetHrs"),
                    "newActualRate": u.get("newActualRate"),
                    "newVariance": u.get("newVariance"),
                }
            result["message"] = "Timesheet update saved for review. If you want to reconcile, then type reconcile."
            result["pending"] = True
        return result
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# --- Get pending updates ---
@app.get("/pending")
def get_pending():
    """Return all pending timesheet updates not yet reconciled to Excel."""
    return {"pending": list(_pending_updates.values()), "count": len(_pending_updates)}


# --- Read employee data ---
@app.get("/employee/{identifier}")
def get_employee(identifier: str):
    try:
        result_json = read_employee_data(identifier=identifier)
        return json.loads(result_json)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# --- Read all data from Excel for the UI ---
@app.get("/data")
def get_all_data():
    try:
        fp = EXCEL_PATH
        if not fp.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {fp}")

        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(list(row))
        wb.close()
        return {"data": rows}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 7. Entrypoint
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    try:
        uvicorn = importlib.import_module("uvicorn")
    except Exception:
        uvicorn = None

    HOST = "0.0.0.0"
    PORT = 8000

    print("=" * 60)
    print("  RECONCILIATION AGENT -- Starting")
    print("=" * 60)

    missing = [v for v in REQUIRED_ENV_VARS if not os.getenv(v)]
    if missing:
        print(f"\n  [FAIL] Missing environment variables: {missing}")
        sys.exit(1)

    try:
        creds = _extract_credentials_from_bearer_token()
        print(f"  [OK] Bearer token valid - expires {creds['expires']}")
        print(f"       Region: {creds['region']}")
    except Exception as e:
        print(f"\n  [FAIL] Credential error: {e}\n")
        sys.exit(1)

    if EXCEL_PATH.exists():
        print(f"  [OK] Excel file found: {EXCEL_PATH}")
    else:
        print(f"  [WARN] Excel file NOT found: {EXCEL_PATH}")

    print()
    print(f"  API URL:  http://localhost:{PORT}")
    print(f"     Health:      GET  http://localhost:{PORT}/health")
    print(f"     Reconcile:   POST http://localhost:{PORT}/reconcile")
    print(f"     Direct:      POST http://localhost:{PORT}/reconcile-direct")
    print(f"     UI-Only:     POST http://localhost:{PORT}/update-ui")
    print(f"     Pending:     GET  http://localhost:{PORT}/pending")
    print(f"     Employee:    GET  http://localhost:{PORT}/employee/{{id_or_name}}")
    print(f"     Data:        GET  http://localhost:{PORT}/data")
    print()
    print("=" * 60)

    uvicorn.run(app, host=HOST, port=PORT)
